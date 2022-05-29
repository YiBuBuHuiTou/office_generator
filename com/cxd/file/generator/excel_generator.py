# -*- coding:UTF-8 -*-
import openpyxl
import sys
import os
import shutil
from data_map import DataMap

message = None
PREFIX = "Floating"
dd = {}


def file_or_dir_judge(src):
    message = "判断输入类型： "
    inputdir = None
    try:
        if os.path.isfile(files):
            message += "输入类型为文件类型"
            inputdir = os.path.dirname(files)

        elif os.path.isdir(files):
            message += "输入类型为文件夹"
            inputdir = files
        else:
            message += "判断失败，请检查输入是否为文件或文件夹"
            raise Exception("Error: 判断失败，请检查输入是否为文件或文件夹")
    except Exception as err:
        sys.exit(1)
    finally:
        print(message)

    return inputdir


def filesBackUp(src):
    outputdir = os.path.join(src, "output")
    message = "判断目标地址是否已存在:"
    if os.path.exists(outputdir):
        message += "目标地址已存在\n进行删除。。。"
        shutil.rmtree(outputdir)
        message += "删除成功"
    print(message)
    message = "开始将源文件或目录拷贝至目标文件夹。。。"
    try:
        shutil.copytree(src, outputdir)
        message += "拷贝成功"
    except Exception as err:
        message += repr(err)
    finally:
        print(message)
    return outputdir


def excel_analyze(src):
    message = "开始解析文件：{}".format(src)
    if not os.path.isfile(src):
        raise Exception("Error: " + src + "非文件类型")
    if os.path.splitext(src)[-1] != ".xls" and os.path.splitext(src)[-1] != ".xlsx":
        raise Exception("Error: " + src + "非excel类型")
    floating_flag = False
    excel = openpyxl.load_workbook(src)
    for worksheet in excel.worksheets:
        for i, row in enumerate(worksheet.rows):
            floating_flag = False
            for j, cell in enumerate(row):
                if cell is not None and cell.value is not None:
                    if cell.value.startswith(PREFIX) and floating_flag is False:
                        cell_raw = cell.value.replace(PREFIX,"")
                        cell.value = cell_replace(cell_raw)
                        print("=================",cell,cell.value)
                        for key, values in dd.items():
                            for k, value in enumerate(values):
                                if k == 0: continue
                                worksheet.insert_rows(i+k+1,1)
                                for col in range(worksheet.max_column-1):
                                  worksheet.cell(i+k+1, col+1).value = cell_replace(cell_raw, key, k)
                                #next(enumerate(worksheet.rows))
                            floating_flag = True
                    else:
                        cell.value = cell_replace(cell.value)
    excel.save(src)
    excel.close()

def cell_copy(src, dst):
    pass



def cell_replace(src,key_v =0, index = 0):
    target = src
    for key, value in dd.items():
        if key == key_v:
            target = target.replace(key, value[index])
        else:
            target = target.replace(key, value[0])
    return target


def file_loop(src):
    if os.path.isfile(src):
        message = "开始解析文件:{}...".format(src)
        excel_analyze(src)
        message += "解析完毕"
        print(message)
        return
    elif os.path.isdir(src):
        for root, dirs, files in os.walk(src):
            for file in files:
                file_loop(os.path.join(root, file))

    else:
        raise Exception("Error: 该文件或目录不存在" + src)


if __name__ == '__main__':
    files = r'F:\python\office_generator\test'
    config_file = r'F:\python\office_generator\新建文本文档.txt'
    message = "开始解析配置。。。"
    dataMap = DataMap(DataMap.TXT, config_file, {})
    PREFIX = dataMap.prefix
    dd = dataMap.convert2dict()
    print(dd)
    message += "解析成功"
    print(message)

    print("读取输入： " + files)
    inputdir = file_or_dir_judge(files)
    outputdir = filesBackUp(inputdir)
    file_loop(outputdir)
